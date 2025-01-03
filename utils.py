import os
import random
import sys
from datetime import datetime
from decimal import Decimal

from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Side, Border, Font, Alignment, PatternFill
from openpyxl.workbook import Workbook

from models import Profile, Result
from settings import max_row_profiles, PROFILES_PATH, RESULTS_PATH, USER_FILES_FOLDER


def get_accounts_from_excel(excel_path: str) -> list:
    profiles = []
    workbook = load_workbook(excel_path)
    sheet = workbook.get_sheet_by_name('not_done')
    # sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=max_row_profiles, min_col=1, max_col=7, values_only=True):
        if not row[1]:
            continue
        print(row)
        profile = Profile(
            id=row[0],
            name=str(row[2]),
            ads_id="".join(char for char in str(row[1]) if not char.isspace()),
            password="".join(char for char in str(row[3]) if not char.isspace()),
            ref_code="".join(char for char in str(row[4]) if not char.isspace()),
        )
        profiles.append(profile)

    logger.info(f"Получил из таблицы profiles листа not_done {len(profiles)} профилей")
    return profiles


def write_results_for_profile(excel_path: str, profile: Profile, result: Result):
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # Создаем стиль для границ (все стороны)
    thin = Side(border_style="thin", color="000000")  # Тонкая черная линия
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    rows = sheet.iter_rows(min_row=2, max_row=max_row_profiles, min_col=1, max_col=9, values_only=True)
    for row_num, row in enumerate(rows, start=2):
        id_ = sheet.cell(row=row_num, column=1, value=profile.id)
        id_.border = border
        ads_id = sheet.cell(row=row_num, column=2, value=profile.ads_id)
        ads_id.border = border
        name = sheet.cell(row=row_num, column=3, value=profile.name)
        name.border = border
        password = sheet.cell(row=row_num, column=4, value=profile.password)
        password.border = border
        ref_code = sheet.cell(row=row_num, column=5, value=profile.ref_code)
        ref_code.border = border


        bubble_amount = sheet.cell(row=row_num, column=6, value=result.bubble_amount)
        bubble_amount.border = border
        tasks_done = sheet.cell(row=row_num, column=7, value=result.tasks_done)
        tasks_done.border = border
        total_win_amount = sheet.cell(row=row_num, column=8, value=result.total_win_amount)
        total_win_amount.border = border
        time = sheet.cell(row=row_num, column=9, value=datetime.now())
        time.border = border

    workbook.save(excel_path)
    workbook.close()


def move_profile_to_done(excel_path: str, profile: Profile):
    workbook = load_workbook(excel_path)
    sheet = workbook.get_sheet_by_name('not_done')
    rows = sheet.iter_rows(min_row=2, max_row=max_row_profiles, min_col=1, max_col=9, values_only=True)
    for i, row in enumerate(rows, start=1):
        if row[1] == profile.ads_id:
            sheet.delete_rows(i)


    sheet = workbook.get_sheet_by_name('done')
    rows = sheet.iter_rows(min_row=2, max_row=max_row_profiles, min_col=1, max_col=9, values_only=True)
    for i, row in enumerate(rows, start=1):
        if row[1] == profile.ads_id:
            workbook.save(excel_path)
            workbook.close()
            return

    row = [profile.id, profile.ads_id, profile.name, profile.password, profile.ref_code]

    sheet.append(row)

    workbook.save(excel_path)
    workbook.close()


def line_control(file_txt):
    # Удаление пустых строк
    with open(file_txt) as f1:
        lines = f1.readlines()
        non_empty_lines = (line for line in lines if not line.isspace())
        with open(file_txt, "w") as n_f1:
            n_f1.writelines(non_empty_lines)


def randfloat(from_: int | float | str, to_: int | float | str,
              step: int | float | str | None = None) -> float:
    """
    Return a random float from the range.

    :param Union[int, float, str] from_: the minimum value
    :param Union[int, float, str] to_: the maximum value
    :param Optional[Union[int, float, str]] step: the step size (calculated based on the number of decimal places)
    :return float: the random float
    """
    from_ = Decimal(str(from_))
    to_ = Decimal(str(to_))
    if not step:
        step = 1 / 10 ** (min(from_.as_tuple().exponent, to_.as_tuple().exponent) * -1)

    step = Decimal(str(step))
    rand_int = Decimal(str(random.randint(0, int((to_ - from_) / step))))
    return float(rand_int * step + from_)

def print_stats(stats: list[Result]):
    for result in stats:
        print(result)


def join_path(path: str | tuple | list) -> str:
    if isinstance(path, str):
        return path
    return str(os.path.join(*path))


def touch(path: str | tuple | list, file: bool = False) -> bool:
    """
    Create an object (file or directory) if it doesn't exist.

    :param Union[str, tuple, list] path: path to the object
    :param bool file: is it a file?
    :return bool: True if the object was created
    """
    path = join_path(path)
    if file:
        if not os.path.exists(path):
            with open(path, 'w') as f:
                f.write('')

            return True

        return False

    if not os.path.isdir(path):
        os.mkdir(path)
        return True

    return False


def create_files():
    touch(USER_FILES_FOLDER)

    # Стиль заголовков
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="FFFF00")  # Желтый фон fgColor="FFFF00"
    header_fill_done = PatternFill("solid", fgColor="FF00FF")  # Желтый фон fgColor="FFFF00"

    # Создаем стиль для границ (все стороны)
    thin = Side(border_style="thin", color="000000")  # Тонкая черная линия
    border = Border(left=thin, right=thin, top=thin, bottom=thin)


    if not os.path.exists(PROFILES_PATH):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "not_done"

        headers = ['id', 'ads_id', 'name', 'Password', 'Ref code']

        # Записываем заголовки
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = border


        sheet_done = workbook.copy_worksheet(sheet)
        sheet_done.title = "done"
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill_done

        workbook.save(PROFILES_PATH)
        workbook.close()


    if not os.path.exists(RESULTS_PATH):
        workbook = Workbook()
        sheet = workbook.active

        headers = ['id', 'ads_id', 'name', 'Password', 'Ref code', 'bubble_amount', 'tasks_done', 'total_win_amount', 'Time']

        # Записываем заголовки
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = border

        workbook.save(RESULTS_PATH)
        workbook.close()


create_files()

logger.remove()
logger.add(
    sys.stdout,
    # colorize=True,
    # format="<light-cyan>{time:HH:mm:ss}</light-cyan> | <level>{level: <8}</level> | <fg #ffffff>{name}:{line}</fg #ffffff> - <bold>{message}</bold>",
)
logger.add(
    './log.log',
    # colorize=True,
    # format="<light-cyan>{time:HH:mm:ss}</light-cyan> | <level>{level: <8}</level> | <fg #ffffff>{name}:{line}</fg #ffffff> - <bold>{message}</bold>",
)
